"""Excel export. Flat data table (no formulas), professionally formatted."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# (header label, record key, width)
COLUMNS = [
    ("Prospect Score", "prospect_score", 13),
    ("Company", "company_name", 26),
    ("Job Title (H1)", "job_title", 28),
    ("Role Offered", "role_bucket", 14),
    ("Roles in Window", "roles_count", 14),
    ("Source Platform", "source_platform", 15),
    ("Posted", "posted_at", 12),
    ("Location", "location", 20),
    ("Domain", "domain", 22),
    ("Industry", "industry", 24),
    ("IT-Native?", "it_native_label", 11),
    ("Employees", "estimated_employees", 11),
    ("Annual Revenue", "annual_revenue_label", 16),
    ("Founded", "founded_year", 9),
    ("Total Funding", "total_funding", 14),
    ("About", "about", 50),
    ("Lead Name", "lead_name", 20),
    ("Lead Title", "lead_title", 22),
    ("Lead LinkedIn", "lead_linkedin", 30),
    ("Lead Email", "lead_email", 26),
    ("Flags", "flags", 34),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")   # charcoal
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SCORE_BANDS = [
    (70, "solid", "C6EFCE"),  # green: strong prospect
    (45, "solid", "FFEB9C"),  # amber: review
    (0, "solid", "F8CBAD"),   # red: weak / likely noise
]


def _score_fill(score):
    for threshold, style, color in SCORE_BANDS:
        if score >= threshold:
            return PatternFill(style, fgColor=color)
    return None


def write_xlsx(records, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # Header
    for ci, (label, _key, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30

    # Rows
    records = sorted(records, key=lambda r: r.get("prospect_score", 0), reverse=True)
    for ri, rec in enumerate(records, start=2):
        for ci, (_label, key, _w) in enumerate(COLUMNS, start=1):
            val = rec.get(key, "")
            if key == "founded_year" and val:
                val = str(val)  # year as text, no thousands separator
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(key in ("about", "flags")))
        sc = rec.get("prospect_score", 0)
        fill = _score_fill(sc)
        if fill:
            ws.cell(row=ri, column=1).fill = fill
            ws.cell(row=ri, column=1).font = Font(name="Arial", size=10, bold=True)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(records) + 1)}"
    wb.save(path)
    return path
